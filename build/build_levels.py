#!/usr/bin/env python3
# ---------------------------------------------------------------------------
# Build the "Levels Program" dataset for Chalk from the Australian Levels
# Program routines workbook.
#
#   python3 build_levels.py  SRC.xlsx  OUT_DIR
#
# Produces, in OUT_DIR:
#   data-levels.js        window.CHALK_LEVELS = { level: { apparatus: [skills] } }
#   levels-images/        the skill pictures, as PNG files referenced by name
#
# WHERE THE DATA LIVES IN THE WORKBOOK
#   One sheet per level ("Level 1" … "Level 8"). Each sheet is a stack of
#   apparatus sections; a section is an ALL-CAPS apparatus header, a note line,
#   a column header row, then the skill rows:
#
#     Skill (number) | Picture | Skill Value | Required Skill | Technical Description | Typical Deductions
#
#   which map to the fields the user asked for:
#     Required Skill        -> name   (what the skill IS)
#     Technical Description -> kcp     (key coaching points, prints in the plan)
#     Picture               -> img     (see below)
#     Skill (number), Skill Value, Typical Deductions -> optional extras
#
# THE PICTURES ARE "IN-CELL" IMAGES, NOT NORMAL DRAWINGS
#   The Picture column uses Excel's rich-value ("Place in Cell") images, which
#   is why the cell reads #VALUE! to anything treating it as text. Resolving one
#   is a five-hop chain through the xlsx parts:
#
#     cell B{row} vm="k"
#        -> metadata.xml  valueMetadata bk[k-1]  rc v=j
#        -> metadata.xml  futureMetadata bk[j]   xlrd:rvb i=r
#        -> richData/rdrichvalue.xml   rv[r]  first <v> = relIndex
#        -> richData/richValueRel.xml  rel[relIndex]  -> rId
#        -> richData/_rels/richValueRel.xml.rels  rId -> ../media/imageN.png
#
#   openpyxl can't read these, so we unzip and walk the XML directly. If a
#   future re-export uses ordinary floating images instead, they'd come through
#   openpyxl's ws._images with a row anchor — not handled here, because this
#   workbook doesn't use them.
# ---------------------------------------------------------------------------
import sys, os, re, json, shutil, zipfile, warnings, base64
warnings.filterwarnings("ignore")
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/MAG_ALP_2026-2029_Routines_Level_1_to_8.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "out"

# Apparatus header text in the sheet  ->  Chalk's canonical apparatus name.
# Chalk already uses these exact names for MAG, so ticking an ALP skill and a
# club skill onto the same rotation just works.
APP_CANON = {
    "FLOOR EXERCISE": "Floor", "FLOOR": "Floor",
    "POMMEL HORSE": "Pommel Horse",
    "RINGS": "Rings",
    "VAULT": "Vault",
    "PARALLEL BARS": "Parallel Bars", "P-BARS": "Parallel Bars",
    "HORIZONTAL BAR": "Horizontal Bar", "HIGH BAR": "Horizontal Bar",
}
HEADER_ROW_MARKERS = {"skill", "picture", "required skill", "technical description"}

def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    # In-cell newlines are just wrapping inside a table cell — collapse to spaces
    # for names/values, but keep them for the multi-point fields (handled below).
    return s.strip()

def is_apparatus_header(ws, r):
    """A section header: the apparatus name in C1, ALL CAPS, and NOT a skill row.

    The Vault sections are the reason for the care here — their skill-number
    cell literally reads "Vault" (rather than "Skill 1"), which is the same word
    as the section header. A real header has no Required Skill beside it and no
    Skill Value; a vault skill row has both. So a "Vault" in C1 with a name in C4
    is a skill, not a new section."""
    c1 = clean(ws.cell(row=r, column=1).value)
    if c1.upper() not in APP_CANON:
        return False
    # If this row also carries a skill name or value, it's a skill row whose
    # number column happens to repeat the apparatus word — not a header.
    if clean(ws.cell(row=r, column=4).value) or clean(ws.cell(row=r, column=3).value):
        return False
    return True

def is_header_row(ws, r):
    vals = [clean(ws.cell(row=r, column=c).value).lower() for c in range(1, 7)]
    return sum(1 for v in vals if v in HEADER_ROW_MARKERS) >= 3

# ---- image resolution -----------------------------------------------------
def build_image_index(xlsx_path):
    """Return { (sheet_index_1based, row) : 'imageN.png' } and the zip handle."""
    z = zipfile.ZipFile(xlsx_path)
    def read(p):
        try: return z.read(p).decode("utf-8", "replace")
        except KeyError: return ""

    # richValueRel index (0-based order of <rel>) -> media filename
    relrels = read("xl/richData/_rels/richValueRel.xml.rels")
    rid_to_media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', relrels))
    rel_order = re.findall(r'<rel r:id="(rId\d+)"', read("xl/richData/richValueRel.xml"))
    idx_to_media = [rid_to_media.get(rid) for rid in rel_order]

    # rv index -> rel index (first <v> in each <rv>)
    rv_xml = read("xl/richData/rdrichvalue.xml")
    rv_to_relidx = []
    for b in re.findall(r'<rv[ >].*?</rv>', rv_xml, re.S):
        vs = re.findall(r'<v>([^<]*)</v>', b)
        rv_to_relidx.append(int(vs[0]) if vs else None)

    # metadata: vm (1-based) -> rc v (j) -> rvb i (r)
    meta = read("xl/metadata.xml")
    rvb = [int(x) for x in re.findall(r'<xlrd:rvb i="(\d+)"', meta)]
    valmeta = re.search(r'<valueMetadata.*?</valueMetadata>', meta, re.S)
    rc_v = re.findall(r'<rc t="1" v="(\d+)"', valmeta.group(0)) if valmeta else []

    def media_for_vm(vm):
        k = int(vm) - 1
        if not (0 <= k < len(rc_v)): return None
        j = int(rc_v[k])
        if not (0 <= j < len(rvb)): return None
        r = rvb[j]
        if not (0 <= r < len(rv_to_relidx)): return None
        ri = rv_to_relidx[r]
        if ri is None or not (0 <= ri < len(idx_to_media)): return None
        return idx_to_media[ri]

    index = {}
    for si in range(1, 9):
        sx = read(f"xl/worksheets/sheet{si}.xml")
        if not sx: continue
        for cm in re.finditer(r'<c r="B(\d+)"[^>]*\bvm="(\d+)"', sx):
            media = media_for_vm(cm.group(2))
            if media:
                index[(si, int(cm.group(1)))] = media
    return index, z

# ---- main -----------------------------------------------------------------
def main():
    img_index, zf = build_image_index(SRC)
    wb = openpyxl.load_workbook(SRC, data_only=True)

    out_img_dir = os.path.join(OUT, "levels-images")
    if os.path.isdir(out_img_dir):
        shutil.rmtree(out_img_dir)
    os.makedirs(out_img_dir, exist_ok=True)

    copied = {}
    b64_index = {}  # "levels-images/alp_imageN.png" -> {d, w, h, e}
    def copy_image(media_name):
        if media_name in copied:
            return copied[media_name]
        data = zf.read(f"xl/media/{media_name}")
        # Give the file a name that says where it belongs, but keep it unique.
        out_name = "alp_" + media_name
        with open(os.path.join(out_img_dir, out_name), "wb") as fh:
            fh.write(data)
        # Also record the bytes + pixel size for the Word-export registry
        # (data-levels-b64.js). The docx builder can't read files from a
        # file:// page, so it needs the picture inlined, exactly like the club
        # diagrams in images-b64.js. Keyed on the SAME "levels-images/…" path
        # the skills carry, so one imageCollector lookup covers both libraries.
        try:
            from PIL import Image
            import io
            w, h = Image.open(io.BytesIO(data)).size
        except Exception:
            w, h = 150, 120
        ext = media_name.rsplit(".", 1)[-1].lower() if "." in media_name else "png"
        b64_index["levels-images/" + out_name] = {
            "d": base64.b64encode(data).decode("ascii"), "w": w, "h": h, "e": ext,
        }
        copied[media_name] = out_name
        return out_name

    levels = {}
    warnings_list = []

    for si, ws in enumerate(wb.worksheets, start=1):
        level_name = ws.title.strip()          # "Level 1" … "Level 8"
        apparatus = {}
        cur_app = None
        max_r, max_c = ws.max_row, ws.max_column

        r = 1
        while r <= max_r:
            if is_apparatus_header(ws, r):
                cur_app = APP_CANON[clean(ws.cell(row=r, column=1).value).upper()]
                apparatus.setdefault(cur_app, [])
                r += 1
                continue
            if cur_app is None or is_header_row(ws, r):
                r += 1
                continue

            name = clean(ws.cell(row=r, column=4).value)   # Required Skill
            kcp_raw = ws.cell(row=r, column=5).value       # Technical Description
            num = clean(ws.cell(row=r, column=1).value)    # "Skill 1" etc
            value = clean(ws.cell(row=r, column=3).value)  # Skill Value
            ded_raw = ws.cell(row=r, column=6).value       # Typical Deductions

            # A skill row is one that names a required skill. Section notes,
            # blank rows and stray sub-headings (e.g. a second "Vault" label)
            # have no name and are skipped.
            if not name:
                r += 1
                continue

            # KCP and deductions wrap mid-sentence in the spreadsheet cell, so a
            # raw line-split gives ragged fragments ("lift arms to a" / "present-
            # ation position"). Rejoin those wrapped fragments back into whole
            # points: a real point break is a line that ends in sentence
            # punctuation, so we keep accumulating until we hit one. Deductions
            # follow a "text 0.1-0.5" pattern, so a value at the end of a line
            # also ends a point.
            def to_points(v, is_deduction=False):
                if v is None:
                    return []
                raw = [ln.strip() for ln in str(v).replace("\r", "\n").split("\n") if ln.strip()]
                points, buf = [], ""
                val_end = re.compile(r'\d(?:\.\d+)?\s*$')          # "... 0.5"
                for ln in raw:
                    buf = (buf + " " + ln).strip() if buf else ln
                    ends = ln.endswith((".", "!", "?", ":", "*"))
                    if is_deduction and val_end.search(ln):
                        ends = True
                    if ends:
                        points.append(buf)
                        buf = ""
                if buf:
                    points.append(buf)
                return points

            skill = {
                "name": name,
                "kcp": to_points(kcp_raw),
            }
            if num:   skill["num"] = num
            if value: skill["value"] = value
            ded = to_points(ded_raw, is_deduction=True)
            if ded:   skill["deductions"] = ded

            media = img_index.get((si, r))
            if media:
                skill["img"] = [copy_image(media)]

            apparatus[cur_app].append(skill)
            r += 1

        # Drop apparatus that ended up empty (header with no skills parsed).
        apparatus = {a: s for a, s in apparatus.items() if s}
        levels[level_name] = {"apparatus": apparatus}

    # Emit the dataset.
    os.makedirs(OUT, exist_ok=True)
    js_path = os.path.join(OUT, "data-levels.js")
    header = (
        "// ==========================================================================\n"
        "// CHALK — AUSTRALIAN LEVELS PROGRAM (MAG, Levels 1-8)  [generated]\n"
        "//\n"
        "// Built from MAG_ALP_2026-2029_Routines_Level_1_to_8.xlsx by\n"
        "// build/build_levels.py. Do not hand-edit — re-run the builder if the\n"
        "// source workbook changes. Coach-made changes belong in the Skills\n"
        "// Library overlay (chalk-library.js), which sits on top of this.\n"
        "//\n"
        "// Shape mirrors CHALK_DATA so the selector treats a level program exactly\n"
        "// like a set of apparatus:\n"
        "//   window.CHALK_LEVELS = {\n"
        "//     'Level 1': { apparatus: { Floor: [ {name, kcp[], img[], num?, value?, deductions?} ] } }\n"
        "//   }\n"
        "// Pictures live as files in levels-images/ and are referenced by name.\n"
        "// ==========================================================================\n"
    )
    with open(js_path, "w", encoding="utf-8") as fh:
        fh.write(header)
        fh.write("window.CHALK_LEVELS = ")
        json.dump(levels, fh, ensure_ascii=False, indent=1)
        fh.write(";\n")

    # Emit the base64 picture registry used by the Word export. Kept in its own
    # file (not merged into the ~2 MB images-b64.js) so the club diagrams can be
    # regenerated independently, and loaded only where the docx export needs it.
    b64_path = os.path.join(OUT, "data-levels-b64.js")
    with open(b64_path, "w", encoding="utf-8") as fh:
        fh.write(
            "// CHALK — Levels Program pictures as base64, for the Word export "
            "[generated].\n"
            "// Same record shape as CHALK_IMG {d,w,h,e}, keyed on the "
            "'levels-images/…' path\n"
            "// each Levels skill carries. Re-run build/build_levels.py to "
            "regenerate.\n"
            "window.CHALK_LEVELS_IMG = "
        )
        json.dump(b64_index, fh, ensure_ascii=False, separators=(",", ":"))
        fh.write(";\n")

    # Report.
    tot_skills = sum(len(s) for lv in levels.values() for s in lv["apparatus"].values())
    tot_imgs = len(copied)
    print(f"levels: {len(levels)}")
    for lv, d in levels.items():
        per = ", ".join(f"{a}:{len(s)}" for a, s in d["apparatus"].items())
        print(f"  {lv}: {per}")
    print(f"total skills: {tot_skills}")
    print(f"images copied: {tot_imgs}")
    print(f"skills with a picture: {sum(1 for lv in levels.values() for s in lv['apparatus'].values() for sk in s if sk.get('img'))}")
    print(f"wrote {js_path} ({os.path.getsize(js_path)//1024} kB)")
    print(f"wrote {b64_path} ({os.path.getsize(b64_path)//1024} kB, {len(b64_index)} pictures)")

if __name__ == "__main__":
    main()
